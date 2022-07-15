import openpyxl as opx
from openpyxl.styles import PatternFill  # , Color, Font, Border

class CommonExcelFileFormatting():
    """Format xlsx file."""

    COLORS = {
        'BISQUE': 'FFE4C4',
        'CYAN': '00FFFF',
        'GRAY': "C0C0C0",
        'LIME': '00FF00',
        'YELLOW': 'FFFF00',
    }

    @staticmethod
    def search_and_colorise(work_sheet, searched_texts_list, color='LIME'):
        """Renewed."""
        if isinstance(searched_texts_list, str):
            raise Exception('list of str expected!')
        # openpx style
        fill_color = PatternFill(start_color=__class__.COLORS[color],
                                 end_color=__class__.COLORS[color],
                                 fill_type='solid')
        for searched in searched_texts_list:
            for j in range(1, work_sheet.max_column + 1):
                # print(work_sheet.cell(row=1, column=j).value)
                if work_sheet.cell(row=1, column=j).value:
                    if work_sheet.cell(row=1, column=j).value.find(searched) >= 0:
                        work_sheet.cell(row=1, column=j).fill = fill_color

    # @staticmethod
    # def header_wrap_and_size(ws, size):
    #     """Renewed."""
    #     for j in range(1, ws.max_column + 1):
    #         ws.cell(row=1, column=j).alignment = \
    #             opx.styles.Alignment(
    #             horizontal='center',  # 'general',
    #             vertical='bottom',
    #             text_rotation=0,
    #             wrap_text=True,
    #             shrink_to_fit=False,
    #             indent=0)
    #         ws.column_dimensions[opx.utils.cell.get_column_letter(j)].width = size
    #         ws.row_dimensions[1].height = 45   # 3 rows height

    @staticmethod
    def start_Excel():
        """Skip."""
        pass  # Excel

    @staticmethod
    def read_xlsxfile(Excel, file_to_format):
        """Skip."""
        return opx.load_workbook(file_to_format)

    @staticmethod
    def read_worksheet(wb, sheet_name: str):
        """Skip."""
        return wb[sheet_name]

    @staticmethod
    def wrap_row(ws, range_: str):
        """Skip."""
        ws.Rows(range_).WrapText = True

    @staticmethod
    def columns_num_format(ws, cols: str):
        """Skip."""
        rows_ = ws.max_row
        cols_from_to_letters = cols.split(':')
        start_index = opx.utils.cell.column_index_from_string(cols_from_to_letters[0])
        end_index = opx.utils.cell.column_index_from_string(cols_from_to_letters[-1])
        for i in range(2, rows_+1):
            for column_index in range(start_index, end_index+1):
                ws.cell(row=i, column=column_index).number_format = '# ### ##0'

    @staticmethod
    def set_columns_width(ws, cols: str, width: float):
        """Set column(s) width."""
        # rows_ = ws.max_row
        cols_from_to_letters = cols.split(':')
        start_index = opx.utils.cell.column_index_from_string(cols_from_to_letters[0])
        end_index = opx.utils.cell.column_index_from_string(cols_from_to_letters[-1])
        for column_index in range(start_index, end_index+1):
            ws.column_dimensions[opx.utils.cell.get_column_letter(column_index)].width = width

    @staticmethod
    def wrap_in_rows(ws, rows: str):
        """Skip."""
        from_to_rows_num = [int(row_num) for row_num in rows.split(':')]
        for row_num in range(from_to_rows_num[0], from_to_rows_num[-1]+1):
            for j in range(1, ws.max_column + 1):
                ws.cell(row=row_num, column=j).alignment = \
                    opx.styles.Alignment(
                    horizontal='center',
                    vertical='bottom',
                    text_rotation=0,
                    wrap_text=True,
                    shrink_to_fit=False,
                    indent=0)
                ws.row_dimensions[row_num].height = 45   # 3 rows height

    @staticmethod
    def columns_autofit(ws, columns: tuple):
        """Autfit column(s) width.

        From 'A' to 'Z' columns only!
        """
        for cols in columns:
            cols_from_to_letters = cols.split(':')
            for ord_ in range(ord(cols_from_to_letters[0]), ord(cols_from_to_letters[-1])+1):
                column_letter = chr(ord_)
                dim = opx.worksheet.dimensions.ColumnDimension(
                    ws, index=column_letter,
                    bestFit=True, customWidth=True,
                )
                ws.column_dimensions[column_letter] = dim

    @staticmethod
    def freeze_panes(Excel, wb, ws, cell: str):
        """Skip."""
        if cell:
            ws.freeze_panes = ws[cell]

    @staticmethod
    def save_workbook(wb, file_to_format):
        """Skip."""
        wb.save(file_to_format)

    @staticmethod
    def close_workbook(wb):
        """Skip."""
        pass

    @staticmethod
    def exit_excel(Excel):
        """Skip."""
        pass

    def __init__(self, formater_dict: dict):
        self.formater_dict = formater_dict


    def excel_file_formatting(self, files_to_format: dict):
        """Format xlsx file."""
        formater = self.formater_dict
        format_xlsx = self

        excel = format_xlsx.start_Excel()
        try:   # for win32com care only ...
            for file_to_format in files_to_format.values():
                wb = format_xlsx.read_xlsxfile(excel, file_to_format)
                try:
                    for page_formater in formater:
                        wsheet = format_xlsx.read_worksheet(wb, page_formater['sheet_name'])
                        # Columns(column_range).AutoFit()
                        if page_formater['columns_autofit']:
                            format_xlsx.columns_autofit(wsheet, page_formater['columns_autofit'])
                        # # freeze panels
                        format_xlsx.freeze_panes(excel, wb, wsheet, page_formater['freeze_panes'])
                        # colorize
                        for color, column_names_list in page_formater['colorize']:
                            format_xlsx.search_and_colorise(wsheet, column_names_list, color=color)
                        # Columns("D:E").NumberFormat = "# ##0"
                        for columns in page_formater['columns_num_format']:
                            format_xlsx.columns_num_format(wsheet, columns)
                        # set columns width
                        for cols, width in page_formater['widths'].items():
                            format_xlsx.set_columns_width(wsheet, cols, width)
                        # Wrap range
                        format_xlsx.wrap_in_rows(wsheet, page_formater['row_wraptext'])
                        format_xlsx.save_workbook(wb, file_to_format)
                finally:
                    format_xlsx.close_workbook(wb)
        finally:
            format_xlsx.exit_excel(excel)